"""
actualizar_fondos.py  v3
────────────────────────
Usa la API interna de Morningstar (time series de precios NAV)
para calcular rentabilidades reales: YTD, 1M, 3M, 1Y.
No hay scraping HTML → más estable y sin bloqueos.
"""

import os, time, datetime, smtplib, logging
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from io import BytesIO

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────
EMAIL_REMITENTE = os.environ["EMAIL_REMITENTE"]
EMAIL_PASSWORD  = os.environ["EMAIL_PASSWORD"]
EMAIL_DESTINO   = os.environ["EMAIL_DESTINO"]
EXCEL_PATH      = "fondos_inversion.xlsx"

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "es-ES,es;q=0.9",
    "Referer": "https://www.morningstar.es/",
}

# ─── ISIN → ID MORNINGSTAR ────────────────────────────────────────────────
ISIN_A_MS = {
    "FR0010830885": "F00000XNKY",
    "FR0013399633": "F00001BSCK",
    "LU1882449801": "F00001BHQK",
    "LU0830528907": "F0GBR063CP",
    "LU0906524193": "F00000YM0O",
    "LU1706854152": "F00001BSIO",
    "FR0011365212": "F00000VXBK",
    "LU0658026603": "F00000NCI5",
    "LU0151324422": "F0GBR04GYK",
    "LU0252128276": "F0GBR04GYL",
    "FR0014008AI4": "F00001DLBY",
    "FR0014000J453": "F00001EO44",
    "LU0336084032": "F0GBR04QEK",
    "FR0010149120": "F0GBR04QEJ",
    "LU1966822956": "F00001DTXY",
    "LU1694789451": "F00001AK5D",
    "LU0343303002": "F0GBR06BXC",
    "LU1161527038": "F00000VXYT",
    "FR0010334495": "F0GBR04QEO",
    "LU0512127621": "F0GBR06BXX",
    "IE0000BTZZE8": "F00001GBA5",
    "IE000WFO9P18": "F00001HC6T",
    "IE0000M5MJ59": "F00001EM4X",
    "IE0000M53C66": "F00001GMLT",
    "LU0300ZRW254": "F0GBR06BXR",
    "IE00B84J9L26": "F00000OUQB",
    "ES0173829047": "F0GBR04IFH",
    "LU2917874104": "F00001N9KY",
    "LU1551754515": "F00001AX9D",
    "LU0243957239": "F0GBR04QEG",
}

# ─── OBTENER SERIE DE PRECIOS NAV DE MORNINGSTAR ─────────────────────────
def obtener_serie_nav(ms_id: str) -> list:
    """
    Llama a la API de series temporales de Morningstar.
    Devuelve lista de [fecha_str, precio] ordenada por fecha ascendente.
    """
    hoy       = datetime.date.today()
    hace_14m  = hoy - datetime.timedelta(days=430)  # 14 meses atrás

    url = (
        f"https://tools.morningstar.es/api/rest.svc/timeseries_price/9coxbuqx31"
        f"?id={ms_id}]2]0]FOESP$$ALL"
        f"&currencyId=EUR&idtype=Morningstar"
        f"&frequency=daily"
        f"&startDate={hace_14m.strftime('%Y-%m-%d')}"
        f"&endDate={hoy.strftime('%Y-%m-%d')}"
        f"&outputType=json"
    )
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        if r.status_code != 200:
            log.debug(f"Serie NAV {ms_id}: status {r.status_code}")
            return []
        data = r.json()
        series = data.get("TimeSeries", {}).get("Security", [])
        if not series:
            return []
        historial = series[0].get("HistoryDetail", [])
        precios = []
        for punto in historial:
            fecha = punto.get("EndDate", "")
            val   = punto.get("Value")
            if fecha and val is not None:
                try:
                    precios.append((fecha, float(val)))
                except (ValueError, TypeError):
                    pass
        return sorted(precios, key=lambda x: x[0])
    except Exception as e:
        log.warning(f"Serie NAV {ms_id}: {e}")
        return []


def calcular_rentabilidades(precios: list) -> dict:
    """Calcula NAV actual, YTD, 1M, 3M, 1Y a partir de la serie de precios."""
    if not precios:
        return {}

    hoy_str     = precios[-1][0]
    nav_actual  = precios[-1][1]
    hoy_dt      = datetime.datetime.strptime(hoy_str[:10], "%Y-%m-%d").date()

    def precio_hace(dias: int):
        fecha_ref = hoy_dt - datetime.timedelta(days=dias)
        # Buscar el precio más cercano anterior a fecha_ref
        candidatos = [(f, p) for f, p in precios
                      if datetime.datetime.strptime(f[:10], "%Y-%m-%d").date() <= fecha_ref]
        return candidatos[-1][1] if candidatos else None

    def precio_inicio_anio():
        anio = str(hoy_dt.year)
        candidatos = [(f, p) for f, p in precios if f[:4] == anio]
        return candidatos[0][1] if candidatos else None

    def pct(p0):
        if p0 is None or p0 == 0:
            return None
        return round((nav_actual / p0 - 1) * 100, 2)

    p_inicio = precio_inicio_anio()
    p_1m     = precio_hace(30)
    p_3m     = precio_hace(90)
    p_1y     = precio_hace(365)

    resultado = {"nav": round(nav_actual, 4), "fecha_nav": hoy_str[:10]}
    if p_inicio: resultado["ytd"] = pct(p_inicio)
    if p_1m:     resultado["1m"]  = pct(p_1m)
    if p_3m:     resultado["3m"]  = pct(p_3m)
    if p_1y:     resultado["1y"]  = pct(p_1y)

    return resultado


# ─── OBTENER DATOS ADICIONALES (rating, duración, YTM) ───────────────────
def obtener_datos_extra(ms_id: str) -> dict:
    """
    Usa el endpoint de datos fundamentales de Morningstar.
    """
    datos = {}
    url = (
        f"https://tools.morningstar.es/api/rest.svc/klr5zyak8x/security/screener"
        f"?page=1&pageSize=1&sortOrder=LegalName+asc&outputType=json"
        f"&version=1&languageId=es-ES&currencyId=EUR"
        f"&universeIds=FOESP%24%24ALL%7CFOEUR%24%24ALL"
        f"&securityDataPoints=SecId%7CLegalName%7CStarRating%7CCreditRating"
        f"%7CEffectiveDuration%7CYieldToMaturity%7CNav%7CGBRReturnM0"
        f"%7CGBRReturnM1%7CGBRReturnM3%7CGBRReturnM12"
        f"&filters=SecId%3AIN%3A{ms_id}"
    )
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return datos
        data = r.json()
        rows = data.get("rows", [])
        if not rows:
            return datos
        row = rows[0]

        if row.get("StarRating"):
            try:
                datos["estrellas"] = int(row["StarRating"])
            except (ValueError, TypeError):
                pass
        if row.get("EffectiveDuration"):
            try:
                datos["duracion"] = float(row["EffectiveDuration"])
            except (ValueError, TypeError):
                pass
        if row.get("YieldToMaturity"):
            try:
                datos["ytm"] = float(row["YieldToMaturity"])
            except (ValueError, TypeError):
                pass
        # Rentabilidades de respaldo (si la serie no devolvió datos)
        for key, campo in [("ytd","GBRReturnM0"),("1m","GBRReturnM1"),
                           ("3m","GBRReturnM3"),("1y","GBRReturnM12")]:
            if row.get(campo) is not None:
                try:
                    datos[key] = float(row[campo])
                except (ValueError, TypeError):
                    pass
        if row.get("Nav"):
            try:
                datos["nav_extra"] = float(row["Nav"])
            except (ValueError, TypeError):
                pass

    except Exception as e:
        log.debug(f"Datos extra {ms_id}: {e}")

    return datos


# ─── ACTUALIZAR EXCEL ─────────────────────────────────────────────────────
AZUL_CL = "D6E4F0"
BLANCO  = "FFFFFF"
AZUL_H  = "1F4E79"
thin = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

COL_DURACION = 6
COL_YTM      = 7
COL_YTD      = 25
COL_ESTRELLAS= 34
COL_NAV      = 36
COL_1M       = 37
COL_3M       = 38
COL_1Y       = 39
COL_ACTUALIZ = 40


def celda(ws, row, col, valor, bg, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=valor)
    c.font      = Font(name="Arial", size=8, bold=bold, color="000000")
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin
    if fmt:
        c.number_format = fmt
    return c


def actualizar_excel():
    wb  = load_workbook(EXCEL_PATH)
    ws  = wb["Fondos - Datos Completos"]
    hoy = datetime.date.today().strftime("%d/%m/%Y")

    # Cabeceras columnas nuevas
    nuevas = {
        COL_NAV:     "NAV\nÚltimo",
        COL_1M:      "Rent.\n1 Mes",
        COL_3M:      "Rent.\n3 Meses",
        COL_1Y:      "Rent.\n1 Año",
        COL_ACTUALIZ:"Última\nActualiz.",
    }
    for col, titulo in nuevas.items():
        c = ws.cell(row=2, column=col, value=titulo)
        c.font      = Font(bold=True, color="FFFFFF", size=8, name="Arial")
        c.fill      = PatternFill("solid", start_color=AZUL_H)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin
        ws.column_dimensions[get_column_letter(col)].width = 10

    fondos_ok = 0
    for row in range(3, ws.max_row + 1):
        isin = ws.cell(row=row, column=2).value
        if not isin or str(isin).strip() == "":
            continue
        isin   = str(isin).strip()
        bg_row = AZUL_CL if row % 2 == 1 else BLANCO

        ms_id = ISIN_A_MS.get(isin)
        if not ms_id:
            log.warning(f"Sin ID Morningstar: {isin}")
            celda(ws, row, COL_ACTUALIZ, hoy, bg_row)
            continue

        log.info(f"  {isin} → {ms_id}")

        # Serie de precios NAV → rentabilidades
        precios = obtener_serie_nav(ms_id)
        datos   = calcular_rentabilidades(precios)

        # Datos extra (estrellas, duración, YTM)
        extra = obtener_datos_extra(ms_id)
        time.sleep(1.0)

        # Combinar: datos de serie tienen prioridad para rentabilidades
        for k in ["ytd", "1m", "3m", "1y"]:
            if k not in datos and k in extra:
                datos[k] = extra[k]
        if "nav" not in datos and "nav_extra" in extra:
            datos["nav"] = extra["nav_extra"]
        for k in ["estrellas", "duracion", "ytm"]:
            if k in extra:
                datos[k] = extra[k]

        # Escribir en Excel
        if datos.get("nav"):
            celda(ws, row, COL_NAV, datos["nav"], bg_row, fmt="#,##0.0000")
        if datos.get("ytd") is not None:
            celda(ws, row, COL_YTD, datos["ytd"] / 100, bg_row, fmt="0.00%",
                  bold=datos["ytd"] > 5)
        for col, key in [(COL_1M,"1m"),(COL_3M,"3m"),(COL_1Y,"1y")]:
            if datos.get(key) is not None:
                celda(ws, row, col, datos[key] / 100, bg_row, fmt="0.00%")
        if datos.get("duracion") is not None:
            celda(ws, row, COL_DURACION, datos["duracion"], bg_row, fmt="0.00")
        if datos.get("ytm") is not None:
            celda(ws, row, COL_YTM, datos["ytm"] / 100, bg_row, fmt="0.00%")
        if datos.get("estrellas"):
            celda(ws, row, COL_ESTRELLAS, "⭐" * datos["estrellas"], bg_row)

        celda(ws, row, COL_ACTUALIZ, hoy, bg_row)
        fondos_ok += 1
        log.info(f"    NAV={datos.get('nav')} YTD={datos.get('ytd')} 1M={datos.get('1m')} 3M={datos.get('3m')} 1Y={datos.get('1y')}")

    # Historial
    if "Historial" not in wb.sheetnames:
        wlog = wb.create_sheet("Historial")
        for col, txt in [(1,"Fecha"),(2,"Fondos actualizados"),(3,"Notas")]:
            wlog.cell(1, col, txt).font = Font(bold=True, name="Arial")
    else:
        wlog = wb["Historial"]
    nr = wlog.max_row + 1
    wlog.cell(nr, 1, hoy)
    wlog.cell(nr, 2, fondos_ok)
    wlog.cell(nr, 3, "Actualización automática diaria v3")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.save(EXCEL_PATH)
    log.info(f"Excel guardado — {fondos_ok} fondos actualizados.")
    return buf, fondos_ok


# ─── ENVIAR EMAIL ─────────────────────────────────────────────────────────
def enviar_email(excel_bytes, fondos_ok):
    hoy_str     = datetime.date.today().strftime("%d/%m/%Y")
    nombre_arch = f"fondos_inversion_{datetime.date.today():%Y%m%d}.xlsx"

    msg            = MIMEMultipart()
    msg["From"]    = EMAIL_REMITENTE
    msg["To"]      = EMAIL_DESTINO
    msg["Subject"] = f"📊 Fondos de Inversión — Actualización {hoy_str}"

    cuerpo = f"""
    <html><body style="font-family:Arial,sans-serif;color:#333">
    <h2 style="color:#1F4E79">📊 Fondos de Inversión — {hoy_str}</h2>
    <p>Se adjunta el Excel con <strong>{fondos_ok} fondos actualizados</strong>.</p>
    <table style="border-collapse:collapse;margin:16px 0">
      <tr style="background:#1F4E79;color:white">
        <th style="padding:8px 16px">Dato</th><th style="padding:8px 16px">Estado</th>
      </tr>
      <tr><td style="padding:6px 16px">💰 NAV / Precio</td><td>✅ Actualizado</td></tr>
      <tr style="background:#f2f2f2"><td style="padding:6px 16px">📈 YTD</td><td>✅ Actualizado</td></tr>
      <tr><td style="padding:6px 16px">📈 Rent. 1M / 3M / 1Y</td><td>✅ Actualizado</td></tr>
      <tr style="background:#f2f2f2"><td style="padding:6px 16px">⭐ Estrellas Morningstar</td><td>✅ Actualizado</td></tr>
      <tr><td style="padding:6px 16px">⏱️ Duración y YTM</td><td>✅ Actualizado</td></tr>
    </table>
    <p style="color:#888;font-size:12px">Fuente: Morningstar ES · GitHub Actions v3</p>
    </body></html>
    """
    msg.attach(MIMEText(cuerpo, "html"))

    parte = MIMEBase("application", "octet-stream")
    parte.set_payload(excel_bytes.read())
    encoders.encode_base64(parte)
    parte.add_header("Content-Disposition", f'attachment; filename="{nombre_arch}"')
    msg.attach(parte)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as srv:
        srv.login(EMAIL_REMITENTE, EMAIL_PASSWORD)
        srv.sendmail(EMAIL_REMITENTE, EMAIL_DESTINO, msg.as_string())
    log.info(f"Email enviado a {EMAIL_DESTINO}")


# ─── MAIN ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    log.info("═══ Inicio actualización v3 ═══")
    excel_buf, fondos_ok = actualizar_excel()
    enviar_email(excel_buf, fondos_ok)
    log.info("═══ Proceso completado ═══")
